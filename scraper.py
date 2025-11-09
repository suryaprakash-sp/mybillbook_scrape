"""
Main scraper for MyBillBook inventory data
"""

import json
import csv
import os
from typing import List, Dict, Any, Optional
from datetime import datetime

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

from auth import MyBillBookAPI
from models import InventoryItem, BulkUploadStatus
from config import OUTPUT_DIR, OUTPUT_FILES


class InventoryScraper:
    """Main class for scraping MyBillBook inventory"""

    def __init__(self, output_dir: str = OUTPUT_DIR, quiet: bool = False):
        """
        Initialize the scraper

        Args:
            output_dir: Directory to save output files
            quiet: If True, suppress non-essential output
        """
        self.api = MyBillBookAPI()
        self.items: List[InventoryItem] = []
        self.all_items: List[InventoryItem] = []  # Store unfiltered items
        self.output_dir = output_dir
        self.quiet = quiet
        self._ensure_output_dir()

    def _ensure_output_dir(self):
        """Create output directory if it doesn't exist"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
            if not self.quiet:
                print(f"Created output directory: {self.output_dir}")

    def fetch_inventory(self) -> bool:
        """
        Fetch all inventory items from MyBillBook

        Returns:
            True if successful, False otherwise
        """
        print("\n" + "=" * 60)
        print("MYBILLBOOK INVENTORY SCRAPER")
        print("=" * 60 + "\n")

        # Test connection first
        if not self.api.test_connection():
            return False

        print("\nFetching complete inventory data...")

        # Get all items from the items API
        response = self.api.get_all_items(per_page=500)

        if not response:
            print("Failed to fetch inventory data.")
            return False

        # Parse the response
        try:
            total_count = response.get('total_count', 0)
            inventory_items = response.get('inventory_items', [])

            print(f"Total items in system: {total_count}")
            print(f"Items fetched: {len(inventory_items)}")

            # Convert API response to our InventoryItem objects
            self.items = []
            for item_data in inventory_items:
                try:
                    # Map the new API structure to our existing model
                    mapped_item = {
                        'id': item_data.get('id'),
                        'name': item_data.get('name'),
                        'sku_code': item_data.get('sku_code', ''),
                        'category': item_data.get('item_category_name', ''),
                        'category_name': item_data.get('item_category_name', ''),
                        'item_category_id': item_data.get('item_category_id', ''),
                        'mrp': float(item_data.get('mrp') or 0),
                        'selling_price': float(item_data.get('selling_price') or 0),
                        'sales_price': float(item_data.get('sales_info', {}).get('price_per_unit') or 0),
                        'purchase_price': float(item_data.get('purchase_price') or 0),
                        'wholesale_price': float(item_data.get('wholesale_info', {}).get('price_per_unit') or 0) if item_data.get('wholesale_info', {}).get('price_per_unit') else None,
                        'wholesale_min_quantity': item_data.get('wholesale_min_quantity'),
                        'quantity': str(item_data.get('quantity', '0')),
                        'minimum_quantity': str(item_data.get('minimum_quantity', '0')),
                        'unit': item_data.get('unit', ''),
                        'unit_long': item_data.get('units', {}).get('primary_unit', item_data.get('unit', '')),
                        'gst_percentage': float(item_data.get('gst_percentage') or 0),
                        'sales_tax_included': item_data.get('is_tax_included', False),
                        'purchase_tax_included': item_data.get('purchase_info', {}).get('is_tax_included', False),
                        'description': item_data.get('description', ''),
                        'item_type': 0,  # Default value since API returns "good" as string
                        'show_on_store': False,  # Not in current API response
                        'excel_imported': False,  # Not in current API response
                        'created_at': item_data.get('created_at', ''),
                        'identification_code': item_data.get('identification_code', ''),
                        'conversion_factor': 0,  # Not in current API response
                        'additional_fields': item_data.get('additional_fields', []),
                        'sub_items': [],  # Not in current API response
                    }

                    item = InventoryItem.from_dict(mapped_item)
                    self.items.append(item)

                except Exception as e:
                    print(f"Warning: Failed to parse item {item_data.get('id', 'unknown')}: {e}")
                    continue

        except Exception as e:
            print(f"Error parsing response: {e}")
            return False

        # Store unfiltered items
        self.all_items = self.items.copy()

        print(f"\n[OK] Successfully fetched {len(self.items)} items!")
        return True

    def apply_filters(
        self,
        category: Optional[str] = None,
        min_stock: Optional[float] = None,
        max_stock: Optional[float] = None,
        min_price: Optional[float] = None,
        max_price: Optional[float] = None
    ):
        """
        Apply filters to the inventory items

        Args:
            category: Filter by category name
            min_stock: Minimum stock quantity
            max_stock: Maximum stock quantity
            min_price: Minimum selling price
            max_price: Maximum selling price
        """
        filtered_items = self.all_items.copy()

        if category:
            filtered_items = [
                item for item in filtered_items
                if category.lower() in (item.category_name or item.category or '').lower()
            ]
            if not self.quiet:
                print(f"Filtered by category '{category}': {len(filtered_items)} items")

        if min_stock is not None:
            filtered_items = [
                item for item in filtered_items
                if float(item.quantity) >= min_stock
            ]
            if not self.quiet:
                print(f"Filtered by min stock {min_stock}: {len(filtered_items)} items")

        if max_stock is not None:
            filtered_items = [
                item for item in filtered_items
                if float(item.quantity) <= max_stock
            ]
            if not self.quiet:
                print(f"Filtered by max stock {max_stock}: {len(filtered_items)} items")

        if min_price is not None:
            filtered_items = [
                item for item in filtered_items
                if item.selling_price >= min_price
            ]
            if not self.quiet:
                print(f"Filtered by min price {min_price}: {len(filtered_items)} items")

        if max_price is not None:
            filtered_items = [
                item for item in filtered_items
                if item.selling_price <= max_price
            ]
            if not self.quiet:
                print(f"Filtered by max price {max_price}: {len(filtered_items)} items")

        self.items = filtered_items

        if not self.quiet:
            print(f"\nTotal items after filters: {len(self.items)}")

    def save_json(self, detailed: bool = False):
        """
        Save inventory data as JSON

        Args:
            detailed: If True, save with full details, else save compact version
        """
        filename = OUTPUT_FILES['detailed_json'] if detailed else OUTPUT_FILES['json']
        filepath = os.path.join(self.output_dir, filename)

        data = [item.to_dict() for item in self.items]

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"[OK] Saved JSON: {filepath} ({len(self.items)} items)")

    def save_csv(self):
        """Save inventory data as CSV"""
        filepath = os.path.join(self.output_dir, OUTPUT_FILES['csv'])

        if not self.items:
            print("No items to save.")
            return

        # Get all field names from the first item
        fieldnames = [
            'id', 'name', 'sku_code', 'category', 'category_name',
            'mrp', 'selling_price', 'sales_price', 'purchase_price',
            'wholesale_price', 'wholesale_min_quantity',
            'quantity', 'minimum_quantity', 'unit', 'unit_long',
            'gst_percentage', 'sales_tax_included', 'purchase_tax_included',
            'description', 'item_type', 'show_on_store', 'excel_imported',
            'created_at', 'identification_code', 'conversion_factor',
            'item_category_id', 'index'
        ]

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for item in self.items:
                row = item.to_dict()
                # Remove complex fields that don't work well in CSV
                row.pop('additional_fields', None)
                row.pop('sub_items', None)
                writer.writerow(row)

        print(f"[OK] Saved CSV: {filepath} ({len(self.items)} items)")

    def save_excel(self):
        """Save inventory data as Excel file"""
        try:
            import pandas as pd
        except ImportError:
            print("Error: pandas and openpyxl are required for Excel export")
            print("Install with: pip install pandas openpyxl")
            return

        filepath = os.path.join(self.output_dir, 'inventory_export.xlsx')

        if not self.items:
            print("No items to save.")
            return

        # Convert items to DataFrame
        data = [item.to_dict() for item in self.items]
        df = pd.DataFrame(data)

        # Remove complex fields
        df = df.drop(columns=['additional_fields', 'sub_items'], errors='ignore')

        # Create Excel writer with auto-adjusting columns
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Inventory')

            # Auto-adjust column widths
            worksheet = writer.sheets['Inventory']
            for idx, column in enumerate(df.columns, 1):
                try:
                    column_width = max(df[column].astype(str).map(len).max(), len(str(column)))
                    # Use get_column_letter for proper Excel column naming
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(idx)
                    worksheet.column_dimensions[col_letter].width = min(column_width + 2, 50)
                except:
                    pass  # Skip if column width adjustment fails

        print(f"[OK] Saved Excel: {filepath} ({len(self.items)} items)")

    def print_summary(self):
        """Print a summary of the scraped inventory"""
        if not self.items:
            print("No items to summarize.")
            return

        print("\n" + "=" * 60)
        print("INVENTORY SUMMARY")
        print("=" * 60)

        # Total items
        print(f"\nTotal Items: {len(self.items)}")

        # Categories
        categories = {}
        for item in self.items:
            cat = item.category_name or item.category
            categories[cat] = categories.get(cat, 0) + 1

        print(f"\nCategories ({len(categories)}):")
        for cat, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
            print(f"  - {cat}: {count} items")

        # Price statistics
        prices = [item.selling_price for item in self.items if item.selling_price > 0]
        if prices:
            print(f"\nPrice Range:")
            print(f"  - Lowest: Rs.{min(prices):,.2f}")
            print(f"  - Highest: Rs.{max(prices):,.2f}")
            print(f"  - Average: Rs.{sum(prices)/len(prices):,.2f}")

        # Total inventory value
        total_value = sum(
            float(item.quantity) * item.selling_price
            for item in self.items
            if item.quantity and item.quantity != 'NaN'
        )
        print(f"\nTotal Inventory Value: Rs.{total_value:,.2f}")

        print("\n" + "=" * 60)

    def export_all(self):
        """Export inventory in all formats"""
        if not self.items:
            print("No items to export.")
            return

        if not self.quiet:
            print("\nExporting inventory data...")
        self.save_json(detailed=False)
        self.save_json(detailed=True)
        self.save_csv()
        self.save_excel()
        if not self.quiet:
            print("\n[OK] All exports completed!")

    def run(self):
        """Main method to run the scraper"""
        # Fetch inventory
        if not self.fetch_inventory():
            print("\n[FAIL] Scraping failed. Please check your configuration and try again.")
            return False

        # Print summary
        self.print_summary()

        # Export data
        self.export_all()

        print("\n" + "=" * 60)
        print("SCRAPING COMPLETED SUCCESSFULLY!")
        print("=" * 60 + "\n")

        return True


def main():
    """Entry point for the scraper"""
    scraper = InventoryScraper()
    success = scraper.run()

    if not success:
        exit(1)


if __name__ == "__main__":
    main()
